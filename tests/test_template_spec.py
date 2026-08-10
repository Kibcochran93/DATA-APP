"""Tests for utils.template_spec (derive a spec from a SEAtS template)."""
import pandas as pd
import pytest

from utils.template_spec import (
    derive_spec_from_template,
    derive_spec_from_workbook,
    derive_spec_from_csv,
    SEATS_GREEN,
)


def _write_template(path, headers, green_indexes):
    """Write an .xlsx whose header row highlights the given columns green."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    green = PatternFill(start_color="FF" + SEATS_GREEN, end_color="FF" + SEATS_GREEN,
                        fill_type="solid")
    wb = Workbook()
    ws = wb.active
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        if (col - 1) in green_indexes:
            cell.fill = green
    wb.save(path)


def test_green_headers_become_mandatory(tmp_path):
    path = tmp_path / "student_template.xlsx"
    headers = ["STUDENT_ID", "STUDENT_FORENAME", "TITLE", "COURSE_ID"]
    _write_template(path, headers, green_indexes={0, 1, 3})  # not TITLE

    spec = derive_spec_from_workbook(path, dataset_type="Student")
    assert spec["dataset_type"] == "Student"
    assert spec["version"] == "template"
    assert spec["mandatory_fields"] == ["STUDENT_ID", "STUDENT_FORENAME", "COURSE_ID"]
    # All headers become fields, in order, with positions.
    assert list(spec["fields"].keys()) == headers
    assert spec["fields"]["STUDENT_ID"]["position"] == 1
    assert spec["fields"]["TITLE"]["mandatory"] is False
    assert spec["fields"]["COURSE_ID"]["mandatory"] is True


def test_workbook_from_bytes(tmp_path):
    path = tmp_path / "t.xlsx"
    _write_template(path, ["EVENT_ID", "DAY"], green_indexes={0, 1})
    spec = derive_spec_from_template(path.read_bytes(), dataset_type="Timetable",
                                     filename="t.xlsx")
    assert spec["mandatory_fields"] == ["EVENT_ID", "DAY"]


def test_no_green_means_all_mandatory(tmp_path):
    path = tmp_path / "plain.xlsx"
    _write_template(path, ["A", "B", "C"], green_indexes=set())
    spec = derive_spec_from_workbook(path, dataset_type="Custom")
    assert spec["mandatory_fields"] == ["A", "B", "C"]


def test_csv_template_all_mandatory(tmp_path):
    path = tmp_path / "cols.csv"
    pd.DataFrame(columns=["STAFF_NUMBER", "FORENAME", "LAST_NAME"]).to_csv(path, index=False)
    spec = derive_spec_from_template(path, dataset_type="Staff")
    assert spec["mandatory_fields"] == ["STAFF_NUMBER", "FORENAME", "LAST_NAME"]
    assert spec["fields"]["FORENAME"]["position"] == 2


def test_badgenumber_blank_header_is_healed(tmp_path):
    path = tmp_path / "quirk.xlsx"
    # The known SEAtS template quirk: blank header between VISAREQUIRED and COURSE_ID.
    _write_template(path, ["STUDENT_ID", "VISAREQUIRED", "", "COURSE_ID"],
                    green_indexes={0})
    spec = derive_spec_from_workbook(path, dataset_type="Student")
    assert "BADGENUMBER" in spec["fields"]


def test_spec_is_accepted_by_seats_validator(tmp_path):
    """A derived spec should plug straight into the runtime validator."""
    from utils.seats_validator import validate_dataset

    path = tmp_path / "mini.xlsx"
    _write_template(path, ["THING_ID", "THING_NAME"], green_indexes={0})
    spec = derive_spec_from_workbook(path, dataset_type="Custom")

    good = pd.DataFrame([{"THING_ID": "1", "THING_NAME": "x"}])
    bad = pd.DataFrame([{"THING_ID": "", "THING_NAME": "x"}])
    assert not [e for e in validate_dataset(good, "Custom", spec=spec).errors
                if e.error_type == "missing_mandatory"]
    assert [e for e in validate_dataset(bad, "Custom", spec=spec).errors
            if e.error_type == "missing_mandatory"]
