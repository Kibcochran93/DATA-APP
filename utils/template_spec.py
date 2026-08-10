"""
Derive a validation spec from a SEAtS template spreadsheet.

SEAtS distributes template workbooks whose header row highlights mandatory
columns with a green fill (RGB 92D050). Instead of hand-authoring a JSON spec
for every dataset type (Room, Activity, Assessment, StudentTags, ...), an
implementer can drop in the template and this derives a spec on the fly:
headings become the field set (in column order) and green-filled headers become
the mandatory fields.

The returned dict is shape-compatible with utils.seats_validator (which accepts
a ``spec=`` argument) and utils.seats_data_handler:

    {
      "dataset_type": <name>,
      "version": "template",
      "source": "template",
      "signature": [first up to 3 headers],
      "mandatory_fields": [green headers],
      "fields": {NAME: {"position": i, "type": "str", "mandatory": bool}, ...},
    }

Ported from the v2.1 browser tool's readFile()/loadTemplate() (which read
SheetJS cellStyles the same way openpyxl exposes cell.fill).
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from utils.integrity_checks import heal_headers

# SEAtS "mandatory" green (openpyxl reports ARGB, e.g. "FF92D050").
SEATS_GREEN = "92D050"


def _rgb_of(cell) -> Optional[str]:
    """Return the 6-digit RGB of a cell's solid fill, or None."""
    fill = getattr(cell, "fill", None)
    if fill is None or getattr(fill, "patternType", None) != "solid":
        return None
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return None
    rgb = getattr(fg, "rgb", None)
    if not isinstance(rgb, str):
        return None
    rgb = rgb.upper()
    if len(rgb) == 8:  # strip ARGB alpha channel
        rgb = rgb[2:]
    return rgb


def _is_green(cell) -> bool:
    rgb = _rgb_of(cell)
    return bool(rgb) and rgb.endswith(SEATS_GREEN)


def _spec_from_headers(headers: List[str], mandatory: List[str], dataset_type: str) -> Dict[str, Any]:
    headers = heal_headers(headers)
    mandatory = [h for h in heal_headers(mandatory) if h]
    # If no colour information was available, treat every column as mandatory
    # (mirrors the v2.1 tool: required = mandatory || headings).
    seen_mandatory = set(mandatory) if mandatory else set(h for h in headers if h)
    fields: Dict[str, Any] = {}
    ordered_mandatory: List[str] = []
    for i, name in enumerate(h for h in headers if h):
        is_mand = name in seen_mandatory
        fields[name] = {"position": i + 1, "type": "str", "mandatory": is_mand}
        if is_mand:
            ordered_mandatory.append(name)
    return {
        "dataset_type": dataset_type,
        "version": "template",
        "source": "template",
        "signature": [h for h in headers if h][:3],
        "mandatory_fields": ordered_mandatory,
        "fields": fields,
    }


def derive_spec_from_workbook(source: Union[str, Path, bytes, io.BytesIO],
                              dataset_type: str = "Custom") -> Dict[str, Any]:
    """Derive a spec from an .xlsx template, using green header fills as mandatory."""
    from openpyxl import load_workbook

    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = load_workbook(source, read_only=False, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        headers: List[str] = []
        mandatory: List[str] = []
        for cell in ws[1]:
            value = "" if cell.value is None else str(cell.value).strip()
            headers.append(value)
            if value and _is_green(cell):
                mandatory.append(value)
    finally:
        wb.close()
    # Trim trailing blank header cells.
    while headers and headers[-1] == "":
        headers.pop()
    return _spec_from_headers(headers, mandatory, dataset_type)


def derive_spec_from_csv(source: Union[str, Path, bytes], dataset_type: str = "Custom") -> Dict[str, Any]:
    """Derive a spec from a CSV template (no colour info -> all columns mandatory)."""
    import pandas as pd

    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    df = pd.read_csv(source, nrows=0)
    headers = [str(c).strip() for c in df.columns]
    return _spec_from_headers(headers, [], dataset_type)


def derive_spec_from_template(source: Union[str, Path, bytes, io.BytesIO],
                              dataset_type: str = "Custom",
                              filename: Optional[str] = None) -> Dict[str, Any]:
    """Derive a spec from a template file (.xlsx/.xls -> fills; .csv -> headers).

    ``filename`` is used to pick the reader when ``source`` is raw bytes.
    """
    name = str(filename or (source if isinstance(source, (str, Path)) else "")).lower()
    if name.endswith(".csv"):
        return derive_spec_from_csv(source, dataset_type)
    return derive_spec_from_workbook(source, dataset_type)
